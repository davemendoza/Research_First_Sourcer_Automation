from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .constants import (
    DEFAULT_SEED_HUB_PATH,
    COL_WATCHLIST_FLAG,
    COL_MONITORING_TIER,
    COL_DOMAIN_TYPE,
    COL_SOURCE_CATEGORY,
    COL_LANGUAGE_CODE,
    DEFAULT_TIER,
    DEFAULT_LANGUAGE_CODE,
)

def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _safe_bool(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"true", "t", "yes", "y", "1"}:
        return True
    if s in {"false", "f", "no", "n", "0"}:
        return False
    return None

def _safe_tier(v: Any) -> str:
    s = _safe_str(v).upper()
    return s if s else DEFAULT_TIER

def _safe_lang(v: Any) -> str:
    s = _safe_str(v).lower()
    return s if s else DEFAULT_LANGUAGE_CODE

@dataclass
class SeedHubRow:
    sheet_name: str
    row_index: int  # 1-based in Excel
    values: Dict[str, Any]

    @property
    def watchlist_flag(self) -> Optional[bool]:
        return _safe_bool(self.values.get(COL_WATCHLIST_FLAG))

    @property
    def monitoring_tier(self) -> str:
        return _safe_tier(self.values.get(COL_MONITORING_TIER))

    @property
    def domain_type(self) -> str:
        return _safe_str(self.values.get(COL_DOMAIN_TYPE))

    @property
    def source_category(self) -> str:
        return _safe_str(self.values.get(COL_SOURCE_CATEGORY))

    @property
    def language_code(self) -> str:
        return _safe_lang(self.values.get(COL_LANGUAGE_CODE))

    def phase_next_metadata(self) -> Dict[str, Any]:
        return {
            "Watchlist_Flag": self.watchlist_flag,
            "Monitoring_Tier": self.monitoring_tier,
            "Domain_Type": self.domain_type,
            "Source_Category": self.source_category,
            "Language_Code": self.language_code,
        }

def _header_map(ws) -> Dict[str, int]:
    header = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = _safe_str(cell.value)
        if key:
            header[key] = idx
    return header

def load_seed_hubs(
    repo_root: Path,
    seed_hub_path: str = DEFAULT_SEED_HUB_PATH,
    max_rows_per_sheet: int = 2500,
) -> List[SeedHubRow]:
    """
    Read Seed Hub workbook rows across all worksheets.
    READ ONLY. No writes, no formatting changes.

    Returns rows that have at least a Primary_URL or Seed Hub Type if present.
    """
    path = (repo_root / seed_hub_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"Seed Hub workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True, keep_links=True)
    rows: List[SeedHubRow] = []

    for ws in wb.worksheets:
        hdr = _header_map(ws)
        if not hdr:
            continue

        # We do not assume exact schema beyond "Primary_URL" or "Seed Hub Type"
        col_primary_url = hdr.get("Primary_URL") or hdr.get("Primary URL") or hdr.get("URL")
        col_seed_type = hdr.get("Seed Hub Type") or hdr.get("Seed_Hub_Type") or hdr.get("Seed_Hub_Type".replace("_", " "))

        # Scan rows
        for r in range(2, min(ws.max_row, max_rows_per_sheet) + 1):
            row_values: Dict[str, Any] = {}
            # Capture only known headers to keep memory sane in read_only mode
            for name, cidx in hdr.items():
                row_values[name] = ws.cell(row=r, column=cidx).value

            primary_url = _safe_str(row_values.get("Primary_URL") or row_values.get("Primary URL") or row_values.get("URL"))
            seed_type = _safe_str(row_values.get("Seed Hub Type") or row_values.get("Seed_Hub_Type") or row_values.get("Seed Hub Type"))

            if not primary_url and not seed_type:
                continue

            rows.append(SeedHubRow(sheet_name=ws.title, row_index=r, values=row_values))

    return rows
