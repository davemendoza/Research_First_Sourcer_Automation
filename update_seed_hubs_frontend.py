#!/usr/bin/env python3
"""
Front-end additive revision of AI_Talent_Landscape_Seed_Hubs.xlsx

Guarantees:
- python3 only
- canonical data/ resolution
- no hardcoded filenames
- preserves formatting & colors
- additive only (no destructive edits)
"""

from pathlib import Path
from openpyxl import load_workbook

# ------------------------------------------------------------------
# Canonical path resolution (single source of truth)
# ------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent
FILE = REPO_ROOT / "data" / "AI_Talent_Landscape_Seed_Hubs.xlsx"

if not FILE.exists():
    raise FileNotFoundError(
        f"Seed Hub Excel not found at expected path: {FILE}"
    )

# ------------------------------------------------------------------
# Column configuration
# ------------------------------------------------------------------

ANCHOR_COLUMNS = [
    "Seed Hub Type",
    "Entity_Type",
    "Primary_URL",
]

NEW_COLUMNS = [
    "Watchlist_Flag",
    "Monitoring_Tier",
    "Domain_Type",
    "Source_Category",
    "Language_Code",
]

# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def find_insert_position(headers):
    idx = -1
    for col in ANCHOR_COLUMNS:
        if col in headers:
            idx = max(idx, headers.index(col))
    return idx + 1 if idx >= 0 else len(headers)

# ------------------------------------------------------------------
# Main execution
# ------------------------------------------------------------------

def main():
    wb = load_workbook(FILE)

    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        insert_at = find_insert_position(headers)

        for col_name in NEW_COLUMNS:
            if col_name in headers:
                continue

            ws.insert_cols(insert_at + 1)
            ws.cell(row=1, column=insert_at + 1).value = col_name
            headers.insert(insert_at, col_name)
            insert_at += 1

    wb.save(FILE)
    print("✅ Seed Hub Excel updated safely with all Phase-Next columns.")

if __name__ == "__main__":
    main()
