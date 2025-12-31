#!/usr/bin/env python3
"""
Seed Hub Loader (Excel → normalized hub dicts)
© 2025 L. David Mendoza

Reads ALL worksheets from the locked Seed Hub workbook (read-only) and produces
normalized hub records for adapter dispatch.

Key normalization:
- Seed_Hub_URL (Excel) → repo_url (adapter contract) for GitHub repo hubs
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


@dataclass(frozen=True)
class HubRecord:
    sheet: str
    row_number: int  # 1-based Excel row number
    data: Dict[str, Any]

    def get(self, key: str, default=None):
        return self.data.get(key, default)


def _normalize_key(k: str) -> str:
    return (k or "").strip()


def load_seed_hubs_excel(xlsx_path: str | Path) -> List[HubRecord]:
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    hubs: List[HubRecord] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue

        if not header:
            continue

        headers = [_normalize_key(h) for h in header]
        if all(h == "" for h in headers):
            continue

        for idx, row in enumerate(rows, start=2):  # row 2 is first data row
            if row is None:
                continue

            record: Dict[str, Any] = {}
            empty = True
            for h, v in zip(headers, row):
                if h == "":
                    continue
                if v is not None and str(v).strip() != "":
                    empty = False
                record[h] = v

            if empty:
                continue

            # Normalization layer (no Excel changes)
            seed_hub_url = record.get("Seed_Hub_URL")
            repo_url = record.get("repo_url") or record.get("Repo_URL") or record.get("repo_url".upper())

            # Canonical fields (normalized)
            record["seed_hub_type"] = record.get("Seed_Hub_Type") or record.get("seed_hub_type")
            record["seed_hub_url"] = seed_hub_url or record.get("seed_hub_url")
            record["repo_url"] = repo_url or seed_hub_url or record.get("seed_hub_url")

            hubs.append(HubRecord(sheet=sheet_name, row_number=idx, data=record))

    return hubs
