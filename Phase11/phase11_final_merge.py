from __future__ import annotations
#!/usr/bin/env python3

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill

# =========================
# CONFIG
# =========================

MODEL_FAMILIES = [
    "GPT", "Claude", "LLaMA", "Mistral", "Mixtral",
    "Gemini", "Qwen", "DeepSeek", "Falcon", "Phi",
    "Gemma", "DBRX", "Grok"
]

TIER_ORDER = {"Tier 1": 1, "Tier 2": 2, "Tier 3": 3, "Tier 4": 4}

CANON_COLUMNS = [
    "Overall Rank",
    "Name",
    "Organization",
    "AI Classification",
    "Technical Signals",
    "Model Families Used",
    "Citations",
    "Citation Velocity",
    "Research Lineage",
    "Strengths",
    "Weaknesses",
    "Tier",
    "Composite Score",
    "Recommendation",
]

# =========================
# UTIL
# =========================

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def s(v):
    return "" if v is None else str(v).strip()

def f(v) -> Optional[float]:
    try:
        return float(v)
    except Exception:
        return None

def parse_series(v) -> List[float]:
    if not v:
        return []
    if isinstance(v, list):
        return [float(x) for x in v if isinstance(x, (int, float))]
    out = []
    for p in str(v).replace(";", ",").split(","):
        try:
            out.append(float(p.strip()))
        except Exception:
            pass
    return out

# =========================
# DATA MODEL
# =========================

@dataclass
class Row:
    name: str
    org: str
    ai: str
    tech: str
    models: str
    citations: str
    velocity: str
    lineage: str
    strengths: str
    weaknesses: str
    tier: str
    score: Optional[float]
    rec: str
    velocity_series: List[float]

    def key(self):
        return (self.name.lower(), self.org.lower())

# =========================
# LOAD
# =========================

def load_candidates(outputs: Path) -> List[Dict[str, Any]]:
    rows = []

    for p in outputs.glob("*.json"):
        if "validation" in p.name or "audit" in p.name:
            continue
        try:
            data = json.loads(p.read_text(errors="replace"))
            rows.extend(data if isinstance(data, list) else [data])
        except Exception:
            pass

    runs = outputs / "runs"
    if runs.exists():
        for p in runs.rglob("*.json"):
            if "candidate" not in p.name.lower():
                continue
            try:
                data = json.loads(p.read_text(errors="replace"))
                rows.extend(data if isinstance(data, list) else [data])
            except Exception:
                pass

    return rows

def normalize(d: Dict[str, Any]) -> Row:
    return Row(
        name=s(d.get("Name")),
        org=s(d.get("Organization")),
        ai=s(d.get("AI Classification")),
        tech=s(d.get("Technical Signals")),
        models=s(d.get("Model Families Used")),
        citations=s(d.get("Citations")),
        velocity=s(d.get("Citation Velocity")),
        lineage=s(d.get("Research Lineage")),
        strengths=s(d.get("Strengths")),
        weaknesses=s(d.get("Weaknesses")),
        tier=s(d.get("Tier")),
        score=f(d.get("Composite Score")),
        rec=s(d.get("Recommendation")),
        velocity_series=parse_series(d.get("Citation_Velocity_Series")),
    )

# =========================
# RANK
# =========================

def rank(rows: List[Row]):
    uniq = {}
    for r in rows:
        if r.key() not in uniq or (r.score or 0) > (uniq[r.key()].score or 0):
            uniq[r.key()] = r

    ordered = sorted(
        uniq.values(),
        key=lambda r: (-(r.score or 0), TIER_ORDER.get(r.tier, 9))
    )
    return [(i + 1, r) for i, r in enumerate(ordered)]

# =========================
# WRITE
# =========================

def write_xlsx(out: Path, ranked):
    wb = Workbook()

    ws = wb.active
    ws.title = "FINAL_FRONTIER_RADAR"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(CANON_COLUMNS)
    for c in range(1, len(CANON_COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    for rank, r in ranked:
        ws.append([
            rank, r.name, r.org, r.ai, r.tech, r.models,
            r.citations, r.velocity, r.lineage,
            r.strengths, r.weaknesses, r.tier,
            r.score, r.rec
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    hm = wb.create_sheet("MODEL_FAMILY_HEATMAP")
    hm.append(["Rank", "Name", "Organization"] + MODEL_FAMILIES)

    for rank, r in ranked:
        text = r.models.lower()
        hm.append([rank, r.name, r.org] + [1 if m.lower() in text else 0 for m in MODEL_FAMILIES])

    end_col = chr(ord("C") + len(MODEL_FAMILIES))
    rule = ColorScaleRule(start_type="num", start_value=0, mid_type="num", mid_value=0.5, end_type="num", end_value=1)
    hm.conditional_formatting.add(f"D2:{end_col}{hm.max_row}", rule)

    meta = wb.create_sheet("META")
    meta.append(["Generated", now()])
    meta.append(["Rows", len(ranked)])

    wb.save(out)

# =========================
# MAIN
# =========================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inputs", default="./outputs")
    ap.add_argument("--out", default="./final_frontier_radar.xlsx")
    args = ap.parse_args()

    raw = load_candidates(Path(args.inputs))
    rows = [normalize(r) for r in raw if r.get("Name")]
    ranked = rank(rows)

    write_xlsx(Path(args.out), ranked)
    print("[ok] wrote", args.out)

if __name__ == "__main__":
    main()
