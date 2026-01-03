"""
Phase-Next Activation (C/D/E) - READ-ONLY

Reads Seed Hub rows, generates:
- Cadence plan (Phase C)
- Watchlist decision (Phase D)
- GPT writeback payload scaffold (Phase E; payload only, no calls)

No writes performed.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from .control_plane import status
from .intelligence_formatter import build_envelope


SEED_HUB_PATH = Path("data/AI_Talent_Landscape_Seed_Hubs.xlsx")


def _load_seed_hub_rows(max_rows: int = 250) -> List[Dict[str, Any]]:
    if not SEED_HUB_PATH.exists():
        raise FileNotFoundError(f"Seed Hub not found at: {SEED_HUB_PATH}")

    wb = load_workbook(SEED_HUB_PATH, data_only=True)
    ws = wb.active

    headers = []
    for c in ws[1]:
        headers.append(str(c.value).strip() if c.value is not None else "")

    rows: List[Dict[str, Any]] = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(rows) >= max_rows:
            break
        row = {}
        for h, v in zip(headers, r):
            if h:
                row[h] = v
        rows.append(row)

    return rows


def main() -> Dict[str, Any]:
    flags = status()

    print("Phase-Next running in READ-ONLY cadence mode")
    print(f"Watchlist enabled: {flags.watchlist_rules_enabled}")
    print(f"GPT writeback enabled: {flags.gpt_writeback_enabled}")
    print(f"Excel writes enabled: {flags.excel_write_enabled}")

    rows = _load_seed_hub_rows(max_rows=250)
    if not rows:
        print("No rows found in Seed Hub.")
        return {"read_only": flags.read_only, "rows_loaded": 0, "writes_performed": False}

    # Build a handful of envelopes as a sanity check
    sample_count = 5 if len(rows) >= 5 else len(rows)
    samples = rows[:sample_count]

    envelopes = [build_envelope(r).to_dict() for r in samples]

    result = {
        "read_only": flags.read_only,
        "excel_write_enabled": flags.excel_write_enabled,
        "watchlist_rules_enabled": flags.watchlist_rules_enabled,
        "gpt_writeback_enabled": flags.gpt_writeback_enabled,
        "seed_hub_path": str(SEED_HUB_PATH),
        "max_rows_total": len(rows),
        "samples_generated": sample_count,
        "writes_performed": False,
        "sample_envelopes": envelopes,
        "notes": "Phase-Next executed in read-only mode. No writes performed.",
    }

    print(json.dumps(result, indent=2))
    print(f"Rows loaded: {len(rows)}")
    print(f"Samples generated: {sample_count}")
    print("Writes performed: False")
    print("Notes: Phase-Next executed in read-only mode. No writes performed.")

    return result


if __name__ == "__main__":
    main()
