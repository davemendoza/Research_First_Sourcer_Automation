#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI Talent Engine – Signal Intelligence
Track D Preflight Validator (Seed Hub Workbook)
Maintainer: Dave Mendoza © 2025
Version: v1.0.0
Status: Production / Fail-Closed
Purpose: Deterministic preflight validation for Seed Hub intelligence architecture template.

───────────────────────────────────────────────────────────────────────────────
WHAT THIS DOES (FAIL-CLOSED)
This validator ensures the Seed Hub workbook is safe and deterministic for Track D.

It enforces:
1) Workbook is READ-ONLY: validator NEVER mutates files.
2) Worksheet schema gate: only worksheets with canonical headers (or declared exceptions) are ingestible.
3) Executable row gate: rows without a Python_Adapter are NON-EXECUTABLE; Track D must not run them.
4) URL integrity: Seed_Hub_URL must be a valid http(s) URL when present; non-URLs are disallowed.
5) Repo contributor readiness: flags any OSS/GitHub hub classes that are not typed for contributor enumeration.
6) Duplicate URL detection: detects duplicated Seed_Hub_URLs within a sheet (risk of duplicate people).
7) Report output: JSON report with errors/warnings and recommended actions.

Exit codes:
- 0: PASS (safe to run Track D)
- 2: FAIL (blocking issues)
- 3: FAIL (report write failure)

───────────────────────────────────────────────────────────────────────────────
LOCKED CONSTRAINTS
- No auto-fixes in this script. This is a validator only.
- No formatting changes. No cell edits. No sheet edits.

───────────────────────────────────────────────────────────────────────────────
CHANGELOG
v1.0.0 (2026-01-25)
- Initial release: schema gate, executable gate, URL integrity, dedupe checks, JSON reporting.

───────────────────────────────────────────────────────────────────────────────
VALIDATION STEPS (RUN THESE)
1) python3 track_d_preflight_validator.py --workbook "/data/AI_Talent_Landscape_Seed_Hubs.xlsx"
2) Confirm "PASS" and exit code 0.
3) If FAIL, open the JSON report and remediate workbook rows (do not guess in Track D).

───────────────────────────────────────────────────────────────────────────────
GIT COMMANDS
git status
git add track_d_preflight_validator.py
git commit -m "Add Track D preflight validator for Seed Hub workbook gates"
git push
───────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


CANONICAL_HEADERS_11 = [
    "Tier",
    "Category",
    "Organization",
    "Seed_Hub_Class",
    "Seed_Hub_Type",
    "Seed_Hub_URL",
    "Primary_Enumeration_Target",
    "Python_Adapter",
    "Expected_Output",
    "Notes",
    "Source",
]

# If you have intentional “non-ingest” sheets, list them here.
# These sheets will still be *reported*, but never treated as ingestible.
NON_INGEST_SHEET_NAMES_DEFAULT = {
    "ReadMe",
    "README",
}

# Minimal, strict URL check. We do not attempt to normalize or “fix” URLs here.
URL_RE = re.compile(r"^https?://", re.IGNORECASE)


@dataclass
class Finding:
    severity: str  # "ERROR" | "WARN" | "INFO"
    code: str
    sheet: str
    row: Optional[int]
    column: Optional[str]
    message: str
    recommendation: Optional[str] = None


@dataclass
class SheetSummary:
    sheet_name: str
    is_ingestible: bool
    header_match: str  # "CANONICAL_11" | "OTHER" | "EMPTY"
    row_count_data: int
    executable_rows: int
    non_executable_rows: int
    url_rows: int
    invalid_url_rows: int
    duplicate_url_count: int


@dataclass
class Report:
    tool: str
    version: str
    timestamp_utc: str
    workbook_path: str
    pass_preflight: bool
    findings: List[Finding]
    sheet_summaries: List[SheetSummary]


def _cell_value(ws: Worksheet, row: int, col: int) -> Optional[str]:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _read_header(ws: Worksheet) -> List[str]:
    # Header is row 1.
    max_col = ws.max_column or 0
    header: List[str] = []
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            header.append("")
        else:
            header.append(str(v).strip())
    # Trim trailing empties
    while header and header[-1] == "":
        header.pop()
    return header


def _header_match_type(header: List[str]) -> str:
    if not header:
        return "EMPTY"
    if header == CANONICAL_HEADERS_11:
        return "CANONICAL_11"
    return "OTHER"


def _build_col_index(header: List[str]) -> Dict[str, int]:
    # 1-based column indices
    return {name: i + 1 for i, name in enumerate(header) if name}


def _iter_data_rows(ws: Worksheet) -> List[int]:
    # Returns row indices that contain any data beyond header.
    rows: List[int] = []
    for r in range(2, ws.max_row + 1):
        # Treat row as data if any cell in the used range is non-empty
        row_has_data = False
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None and str(ws.cell(row=r, column=c).value).strip() != "":
                row_has_data = True
                break
        if row_has_data:
            rows.append(r)
    return rows


def _is_valid_url(url: str) -> bool:
    return bool(URL_RE.match(url))


def validate_workbook(
    workbook_path: Path,
    non_ingest_sheet_names: set[str],
    fail_on_any_noncanonical_sheet: bool,
    fail_on_duplicates: bool,
) -> Report:
    findings: List[Finding] = []
    summaries: List[SheetSummary] = []

    if not workbook_path.exists():
        findings.append(
            Finding(
                severity="ERROR",
                code="WORKBOOK_NOT_FOUND",
                sheet="(workbook)",
                row=None,
                column=None,
                message=f"Workbook not found: {workbook_path}",
                recommendation="Fix path and rerun preflight.",
            )
        )
        return Report(
            tool="track_d_preflight_validator",
            version="v1.0.0",
            timestamp_utc=datetime.utcnow().isoformat() + "Z",
            workbook_path=str(workbook_path),
            pass_preflight=False,
            findings=findings,
            sheet_summaries=summaries,
        )

    # Read-only load. No write operations performed.
    wb = load_workbook(filename=str(workbook_path), data_only=True, read_only=True)

    pass_preflight = True

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = _read_header(ws)
        header_type = _header_match_type(header)

        is_non_ingest = sheet_name in non_ingest_sheet_names
        is_ingestible = (not is_non_ingest) and (header_type == "CANONICAL_11")

        if header_type == "EMPTY":
            # Empty sheets are not a failure, but they are non-ingestible.
            findings.append(
                Finding(
                    severity="INFO",
                    code="SHEET_EMPTY",
                    sheet=sheet_name,
                    row=None,
                    column=None,
                    message="Sheet appears empty or missing headers.",
                    recommendation="No action required unless this sheet was expected to be ingestible.",
                )
            )
        elif header_type == "OTHER":
            findings.append(
                Finding(
                    severity="WARN" if not fail_on_any_noncanonical_sheet else "ERROR",
                    code="SHEET_NONCANONICAL_HEADER",
                    sheet=sheet_name,
                    row=1,
                    column=None,
                    message=f"Header does not match canonical 11-column schema. Found {len(header)} columns.",
                    recommendation=(
                        "Either (1) exclude this sheet from Track D ingestion, or "
                        "(2) normalize headers to canonical schema if intended for ingestion."
                    ),
                )
            )
            if fail_on_any_noncanonical_sheet and not is_non_ingest:
                pass_preflight = False

        if is_non_ingest:
            findings.append(
                Finding(
                    severity="INFO",
                    code="SHEET_MARKED_NON_INGEST",
                    sheet=sheet_name,
                    row=None,
                    column=None,
                    message="Sheet is marked as NON-INGEST by configuration.",
                    recommendation="Track D must skip this sheet explicitly.",
                )
            )

        data_rows = _iter_data_rows(ws)

        executable_rows = 0
        non_exec_rows = 0
        url_rows = 0
        invalid_url_rows = 0

        duplicate_url_count = 0
        seen_urls: Dict[str, int] = {}

        if header_type == "CANONICAL_11":
            col = _build_col_index(header)

            # Validate data rows
            for r in data_rows:
                url = _cell_value(ws, r, col["Seed_Hub_URL"])
                adapter = _cell_value(ws, r, col["Python_Adapter"])
                hub_class = _cell_value(ws, r, col["Seed_Hub_Class"])
                hub_type = _cell_value(ws, r, col["Seed_Hub_Type"])
                notes = _cell_value(ws, r, col["Notes"])

                # Executable gate
                if adapter is None or adapter == "" or adapter.upper() == "N/A":
                    non_exec_rows += 1
                    # This is a blocking issue only if the row also has a URL (i.e., looks runnable but lacks adapter)
                    if url:
                        findings.append(
                            Finding(
                                severity="ERROR",
                                code="ROW_NO_ADAPTER",
                                sheet=sheet_name,
                                row=r,
                                column="Python_Adapter",
                                message="Row has a Seed_Hub_URL but no Python_Adapter. This row must be non-executable.",
                                recommendation="Set a valid Python_Adapter or remove/blank the Seed_Hub_URL.",
                            )
                        )
                        pass_preflight = False
                else:
                    executable_rows += 1

                # URL integrity
                if url:
                    url_rows += 1
                    if not _is_valid_url(url):
                        invalid_url_rows += 1
                        findings.append(
                            Finding(
                                severity="ERROR",
                                code="INVALID_URL",
                                sheet=sheet_name,
                                row=r,
                                column="Seed_Hub_URL",
                                message=f"Seed_Hub_URL is not a valid http(s) URL: {url!r}",
                                recommendation="Replace with a valid https:// URL or blank the field.",
                            )
                        )
                        pass_preflight = False
                    else:
                        # Duplicate URL detection (within this sheet)
                        k = url.lower()
                        if k in seen_urls:
                            duplicate_url_count += 1
                            findings.append(
                                Finding(
                                    severity="WARN" if not fail_on_duplicates else "ERROR",
                                    code="DUPLICATE_URL",
                                    sheet=sheet_name,
                                    row=r,
                                    column="Seed_Hub_URL",
                                    message=f"Duplicate URL detected. First seen at row {seen_urls[k]}.",
                                    recommendation="Canonicalize to a single row per URL, or enforce runtime dedupe in Track D.",
                                )
                            )
                            if fail_on_duplicates:
                                pass_preflight = False
                        else:
                            seen_urls[k] = r

                # Repo contributor readiness (strict, informational gate)
                # If hub_class indicates OSS/GitHub and hub_type is not GitHub_Repo_Contributors, warn.
                if hub_class and hub_class.strip().lower() in {"oss / github", "oss/github", "oss - github", "oss"}:
                    if hub_type and hub_type != "GitHub_Repo_Contributors":
                        findings.append(
                            Finding(
                                severity="WARN",
                                code="OSS_TYPE_NOT_CONTRIB",
                                sheet=sheet_name,
                                row=r,
                                column="Seed_Hub_Type",
                                message=(
                                    "OSS/GitHub hub class row is not typed as GitHub_Repo_Contributors. "
                                    "This may prevent deterministic contributor enumeration."
                                ),
                                recommendation="Set Seed_Hub_Type=GitHub_Repo_Contributors for contributor-enumerator repos.",
                            )
                        )

                # Notes marker check (informational)
                if notes and "NON_EXECUTABLE_NO_ADAPTER" in notes and adapter and adapter.upper() != "N/A":
                    findings.append(
                        Finding(
                            severity="INFO",
                            code="NOTES_CONTRADICT_ADAPTER",
                            sheet=sheet_name,
                            row=r,
                            column="Notes",
                            message="Notes contains NON_EXECUTABLE_NO_ADAPTER but Python_Adapter is present.",
                            recommendation="Remove the NON_EXECUTABLE marker if this row is now executable.",
                        )
                    )

        # Summarize
        summaries.append(
            SheetSummary(
                sheet_name=sheet_name,
                is_ingestible=is_ingestible,
                header_match=header_type,
                row_count_data=len(data_rows),
                executable_rows=executable_rows,
                non_executable_rows=non_exec_rows,
                url_rows=url_rows,
                invalid_url_rows=invalid_url_rows,
                duplicate_url_count=duplicate_url_count,
            )
        )

    # If any ERROR finding exists, fail.
    for f in findings:
        if f.severity == "ERROR":
            pass_preflight = False
            break

    return Report(
        tool="track_d_preflight_validator",
        version="v1.0.0",
        timestamp_utc=datetime.utcnow().isoformat() + "Z",
        workbook_path=str(workbook_path),
        pass_preflight=pass_preflight,
        findings=findings,
        sheet_summaries=summaries,
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="track_d_preflight_validator",
        description="Fail-closed preflight validator for Seed Hub workbook before Track D execution.",
    )
    parser.add_argument(
        "--workbook",
        required=True,
        help="Path to Seed Hub workbook (read-only). Example: /data/AI_Talent_Landscape_Seed_Hubs.xlsx",
    )
    parser.add_argument(
        "--report",
        default="track_d_preflight_report.json",
        help="Path to write JSON report (default: track_d_preflight_report.json).",
    )
    parser.add_argument(
        "--fail-on-noncanonical-sheets",
        action="store_true",
        help="Fail if any non-ReadMe sheet does not match canonical headers.",
    )
    parser.add_argument(
        "--fail-on-duplicates",
        action="store_true",
        help="Fail if duplicate Seed_Hub_URLs are detected within ingestible sheets.",
    )
    parser.add_argument(
        "--non-ingest-sheets",
        default=",".join(sorted(NON_INGEST_SHEET_NAMES_DEFAULT)),
        help="Comma-separated list of sheet names to treat as NON-INGEST (default includes ReadMe).",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    report_path = Path(args.report).expanduser()

    non_ingest = set([s.strip() for s in args.non_ingest_sheets.split(",") if s.strip()])

    report = validate_workbook(
        workbook_path=workbook_path,
        non_ingest_sheet_names=non_ingest,
        fail_on_any_noncanonical_sheet=bool(args.fail_on_noncanonical_sheets),
        fail_on_duplicates=bool(args.fail_on_duplicates),
    )

    try:
        payload = asdict(report)
        # Convert dataclasses in lists
        payload["findings"] = [asdict(f) for f in report.findings]
        payload["sheet_summaries"] = [asdict(s) for s in report.sheet_summaries]
        report_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"[FAIL] Could not write report to {report_path}: {e}", file=sys.stderr)
        return 3

    if report.pass_preflight:
        print("[PASS] Track D preflight OK. Safe to run Track D.")
        print(f"[OK] Report written: {report_path}")
        return 0

    print("[FAIL] Track D preflight FAILED. Do not run Track D until fixed.")
    print(f"[INFO] Report written: {report_path}")
    # Print the first few blocking errors for convenience
    blocking = [f for f in report.findings if f.severity == "ERROR"]
    for f in blocking[:15]:
        loc = f"{f.sheet}"
        if f.row:
            loc += f":R{f.row}"
        if f.column:
            loc += f":{f.column}"
        print(f"  - {loc} [{f.code}] {f.message}")
    if len(blocking) > 15:
        print(f"  ... plus {len(blocking)-15} more ERROR findings.")
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
