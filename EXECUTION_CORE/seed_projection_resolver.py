# -*- coding: utf-8 -*-
"""
seed_projection_resolver.py
------------------------------------------------------------
IMPORT-ONLY MODULE (READ-ONLY SEED HUB XLSX -> CSV PROJECTION)

Maintainer: L. David Mendoza © 2026
Version: v1.1.0-master-seed-hubs-projection-locked

Purpose
- Your seed hub workbook does NOT contain role-named tabs.
- It contains MASTER_SEED_HUBS (and Tier sheets) that define seed SOURCE HUBS.
- This resolver deterministically projects a runnable seed CSV for any role key by
  filtering MASTER_SEED_HUBS into repo_root/seeds/{seed_key}.csv.

Hard Contract
- Read-only: never modifies the Excel hub.
- Deterministic: stable ordering, stable selection.
- Non-destructive: never overwrites an existing seed CSV.
- No network calls.
- Single responsibility: projection only.

Hub reality (validated)
- Workbook contains: MASTER_SEED_HUBS with headers including:
  Tier, Category, Organization, Seed_Hub_URL, etc.
- This module uses MASTER_SEED_HUBS as source of truth.

Output CSV
- Created only if missing.
- Always includes a canonical "URL" column populated from Seed_Hub_URL so downstream
  URL exhaustion and identity passes can see it.
- Includes all MASTER_SEED_HUBS columns (excluding empty header cells).

Validation Steps
python3 -m py_compile EXECUTION_CORE/seed_projection_resolver.py
python3 -c "from EXECUTION_CORE.seed_projection_resolver import ensure_seed_csv_for_seed_key; \
from pathlib import Path; \
print(ensure_seed_csv_for_seed_key('frontier_ai_scientist', Path('.')))"

Git Commands
git add EXECUTION_CORE/seed_projection_resolver.py
git commit -m "Seed resolver: project runnable role seeds from MASTER_SEED_HUBS"
git push
"""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# Known hub filenames. The hub may live in repo root or repo_root/data.
KNOWN_HUB_FILENAMES: Tuple[str, ...] = (
    "AI_Talent_Landscape_Seed_Hubs.xlsx",
    "AI_Talent_Engine_Demo_CANONICAL_81cols_WORKBOOK_v4.xlsx",
)

# Source sheet in your hub workbook
MASTER_SHEET_NAME = "MASTER_SEED_HUBS"

# Deterministic tier ordering
_TIER_ORDER: Dict[str, int] = {
    "Tier 1": 1,
    "Tier1": 1,
    "Tier 2": 2,
    "Tier2": 2,
    "Tier 3": 3,
    "Tier3": 3,
    "Cross-Tier": 4,
    "CrossTier": 4,
    "Cross Tier": 4,
}

# Categories present in your hub (validated from workbook)
# Academic & Institutes
# Ecosystem & Fast Movers
# International & Chinese AI Labs
# Major Tech Labs
# Primary AGI Research Labs
# (and priority reference categories like Conferences, Awards, Patents)
#
# Projection is intentionally conservative: it targets the major seed hub categories.
_ROLE_FILTERS: Dict[str, Dict[str, List[str]]] = {
    # Core Research & Frontier
    "frontier_ai_scientist": {
        "tiers": ["Tier 1", "Tier 2"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Academic & Institutes",
            "International & Chinese AI Labs",
        ],
    },
    "foundational_ai_scientist": {
        "tiers": ["Tier 1", "Tier 2"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Academic & Institutes",
            "International & Chinese AI Labs",
        ],
    },
    "ai_research_scientist": {
        "tiers": ["Tier 1", "Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Academic & Institutes",
            "International & Chinese AI Labs",
            "Primary AGI Research Labs",
        ],
    },
    "machine_learning_researcher": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Academic & Institutes",
            "Major Tech Labs",
        ],
    },

    # Modeling, Training & RL
    "machine_learning_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "applied_machine_learning_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Ecosystem & Fast Movers",
            "Major Tech Labs",
        ],
    },
    "rlhf_engineer": {
        "tiers": ["Tier 1", "Tier 2"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
        ],
    },
    "reinforcement_learning_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Academic & Institutes",
        ],
    },
    "model_training_engineer": {
        "tiers": ["Tier 1", "Tier 2", "Tier 3"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },

    # Infrastructure, Systems & Platform
    "ai_infrastructure_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "machine_learning_infrastructure_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "distributed_systems_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "site_reliability_engineer_for_ai": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "gpu_systems_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },

    # Deployment & customer roles (still driven by hub categories)
    "ai_solutions_architect": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "forward_deployed_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Ecosystem & Fast Movers",
            "Major Tech Labs",
        ],
    },
    "ai_platform_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },

    # Safety, evaluation, alignment
    "ai_safety_engineer": {
        "tiers": ["Tier 1", "Tier 2", "Tier 3"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Academic & Institutes",
        ],
    },
    "model_evaluation_engineer": {
        "tiers": ["Tier 2", "Tier 3"],
        "categories": [
            "Major Tech Labs",
            "Ecosystem & Fast Movers",
        ],
    },
    "ai_alignment_engineer": {
        "tiers": ["Tier 1", "Tier 2"],
        "categories": [
            "Primary AGI Research Labs",
            "Major Tech Labs",
        ],
    },
}


def resolve_hub_path(repo_root: Path) -> Path:
    """
    Determine the authoritative hub workbook path.

    Order:
    1) AI_TALENT_SEED_HUB env var (absolute or relative to repo root)
    2) repo root for known filenames
    3) repo_root/data for known filenames
    """
    env = (os.getenv("AI_TALENT_SEED_HUB") or "").strip()
    if env:
        p = Path(env).expanduser()
        if not p.is_absolute():
            p = (repo_root / p).resolve()
        if p.exists() and p.is_file():
            return p
        raise FileNotFoundError(f"AI_TALENT_SEED_HUB points to missing file: {p}")

    for base in (repo_root, repo_root / "data"):
        for name in KNOWN_HUB_FILENAMES:
            p = base / name
            if p.exists() and p.is_file():
                return p

    tried = []
    for base in (repo_root, repo_root / "data"):
        for name in KNOWN_HUB_FILENAMES:
            tried.append(str(base / name))

    raise FileNotFoundError(
        "No seed hub workbook found.\n"
        "Checked:\n  - " + "\n  - ".join(tried)
    )


def _normalize_header(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _tier_rank(t: str) -> int:
    return _TIER_ORDER.get((t or "").strip(), 999)


def _safe_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _read_master_seed_hubs(ws) -> Tuple[List[str], List[Dict[str, str]]]:
    """
    Read MASTER_SEED_HUBS into (headers, rows).
    Drops trailing empty header cells.
    """
    max_col = ws.max_column or 0
    raw_headers = [_normalize_header(ws.cell(row=1, column=c).value) for c in range(1, max_col + 1)]

    while raw_headers and raw_headers[-1] == "":
        raw_headers.pop()

    headers = [h for h in raw_headers if h != ""]
    raw_count = len(raw_headers)

    rows: List[Dict[str, str]] = []
    started = False

    for r in range(2, (ws.max_row or 0) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, raw_count + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            if started:
                break
            continue

        started = True
        d: Dict[str, str] = {}
        for idx, h in enumerate(raw_headers):
            if h == "":
                continue
            d[h] = _safe_str(vals[idx] if idx < len(vals) else "")
        rows.append(d)

    return headers, rows


def _project_rows_for_seed_key(seed_key: str, rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Deterministic projection from MASTER_SEED_HUBS into a role seed list.
    If role filter is unknown, fall back to a safe default that still enables demos.
    """
    filt = _ROLE_FILTERS.get(seed_key)

    if not filt:
        # Safe default: enable any-role demo by returning Tier 1-3 from core categories.
        tiers = {"Tier 1", "Tier 2", "Tier 3"}
        cats = {
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Academic & Institutes",
            "International & Chinese AI Labs",
            "Ecosystem & Fast Movers",
        }
    else:
        tiers = set(filt.get("tiers", []))
        cats = set(filt.get("categories", []))

    out = []
    for r in rows:
        tier = r.get("Tier", "")
        cat = r.get("Category", "")
        if tier in tiers and cat in cats:
            out.append(r)

    # If projection returns nothing, fail open to a minimal deterministic set
    # to keep demos runnable, but do not fabricate.
    if not out:
        out = [r for r in rows if r.get("Tier", "") in {"Tier 1", "Tier 2"} and r.get("Category", "") in {
            "Primary AGI Research Labs",
            "Major Tech Labs",
            "Academic & Institutes",
        }]

    # Deterministic ordering
    def sort_key(d: Dict[str, str]):
        return (
            _tier_rank(d.get("Tier", "")),
            (d.get("Category", "") or "").lower(),
            (d.get("Organization", "") or "").lower(),
            (d.get("Seed_Hub_URL", "") or "").lower(),
            (d.get("Seed_Hub_Type", "") or "").lower(),
        )

    out = sorted(out, key=sort_key)

    # Deterministic de-duplication by Seed_Hub_URL
    seen = set()
    deduped = []
    for d in out:
        u = (d.get("Seed_Hub_URL", "") or "").strip()
        k = u.lower()
        if not u or k in seen:
            continue
        seen.add(k)
        deduped.append(d)

    return deduped


def _candidate_existing_seed_csv_paths(repo_root: Path, seed_key: str) -> List[Path]:
    """
    Mirrors run_safe.py naming patterns for compatibility.
    """
    names = [
        f"{seed_key}.csv",
        f"{seed_key}_seed.csv",
        f"seed_{seed_key}.csv",
    ]
    dirs = [
        repo_root,
        repo_root / "inputs",
        repo_root / "data",
        repo_root / "seeds",
    ]
    out: List[Path] = []
    for d in dirs:
        for n in names:
            out.append(d / n)
    return out


def ensure_seed_csv_for_seed_key(seed_key: str, repo_root: Path) -> Optional[Path]:
    """
    Ensure an executable CSV seed exists for seed_key.

    Returns:
      - Path to existing seed CSV if present
      - Path to newly projected seed CSV if created
      - None only if hub cannot be read (missing workbook or missing MASTER_SEED_HUBS)

    Non-overwrite guarantee:
      - If seeds/{seed_key}.csv already exists, it is returned unchanged.
    """
    # 1) If any acceptable seed CSV exists, do nothing.
    candidates = _candidate_existing_seed_csv_paths(repo_root, seed_key)
    for p in candidates:
        if p.exists() and p.is_file():
            return p

    # 2) Project from MASTER_SEED_HUBS into repo_root/seeds/{seed_key}.csv
    hub = resolve_hub_path(repo_root)

    wb = load_workbook(filename=str(hub), read_only=True, data_only=True)
    try:
        if MASTER_SHEET_NAME not in wb.sheetnames:
            return None

        ws = wb[MASTER_SHEET_NAME]
        headers, master_rows = _read_master_seed_hubs(ws)
        projected_rows = _project_rows_for_seed_key(seed_key, master_rows)
        if not projected_rows:
            return None

        out_dir = repo_root / "seeds"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{seed_key}.csv"

        if out_path.exists():
            return out_path

        # Output headers: canonical URL first, then the master headers (excluding empty)
        out_headers = ["URL"] + [h for h in headers if h != ""]

        with out_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=out_headers)
            w.writeheader()
            for r in projected_rows:
                row_out = dict(r)
                row_out["URL"] = (r.get("Seed_Hub_URL", "") or "").strip()
                w.writerow({k: row_out.get(k, "") for k in out_headers})

        return out_path
    finally:
        try:
            wb.close()
        except Exception:
            pass
